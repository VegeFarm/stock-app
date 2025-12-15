import io
import json
import re
from copy import copy
from datetime import datetime

import pandas as pd
import streamlit as st
from PIL import Image

import openpyxl
import pytesseract

try:
    import cv2
    import numpy as np
except Exception:
    cv2 = None
    np = None


# =========================
# 설정
# =========================
CATEGORY_TO_COLUMN = {
    "재고": 8,   # H
    "입고": 9,   # I
    "1차": 11,   # K
    "2차": 12,   # L
    "3차": 13,   # M
}
PRODUCT_COL = 7  # G (품목명)
HEADER_ROW = 1
FIRST_DATA_ROW = 2

MAPPING_FILE = "product_mapping.json"


# =========================
# 유틸
# =========================
def load_mapping() -> dict:
    try:
        with open(MAPPING_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_mapping(mapping: dict) -> None:
    with open(MAPPING_FILE, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)


def normalize_name(s: str) -> str:
    if not s:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s+", "", s)
    return s


def parse_qty(qty_text: str) -> float:
    """
    예: '31BOX*1' / '11KG * 1' / '3*2' / '4.6KG×1'
    -> 숫자들 뽑아서 앞*뒤(있으면) 처리
    """
    if not qty_text:
        return 0.0
    t = str(qty_text).upper().replace("×", "*").replace("X", "*")
    nums = re.findall(r"(\d+(?:\.\d+)?)", t)
    if not nums:
        return 0.0
    if len(nums) >= 2:
        return float(nums[0]) * float(nums[1])
    return float(nums[0])


def preprocess_for_ocr(pil_img: Image.Image) -> Image.Image:
    """
    OCR 정확도를 조금 올리기 위한 간단 전처리(가능할 때만)
    """
    if cv2 is None or np is None:
        return pil_img

    img = np.array(pil_img.convert("RGB"))
    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    gray = cv2.bilateralFilter(gray, 9, 75, 75)
    _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return Image.fromarray(th)


def ocr_extract_text(pil_img: Image.Image, tesseract_cmd: str | None) -> str:
    """
    Tesseract OCR
    """
    if tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

    img = preprocess_for_ocr(pil_img)

    # 한국어+영어 우선, 실패하면 영어만
    try:
        return pytesseract.image_to_string(img, lang="kor+eng")
    except Exception:
        return pytesseract.image_to_string(img, lang="eng")


def parse_table_text_to_items(text: str) -> pd.DataFrame:
    """
    OCR로 나온 텍스트에서 (품목, 단위/수량) 중심으로 최대한 뽑아냄.
    스샷 형태(번호. 품목  31BOX*1  4,800  148,800)를 가정.
    """
    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue

        # "1. 적근대 31BOX*1 4,800 148,800" 같은 줄 시도
        m = re.match(r"^\s*\d+\.\s*(.+)$", line)
        if not m:
            continue

        rest = m.group(1).strip()

        # 품목명: 앞쪽 한글/영문 덩어리
        # 수량: 중간에 00BOX*1 / 00KG*1 같은 패턴을 우선 탐색
        qty_m = re.search(r"(\d+(?:\.\d+)?\s*[A-Z가-힣]+?\s*[*X×]\s*\d+(?:\.\d+)?)", rest, re.IGNORECASE)
        if qty_m:
            qty_text = qty_m.group(1)
            name_part = rest[: qty_m.start()].strip()
        else:
            # 수량 패턴이 못 잡히면 숫자+단위라도 잡아보기 (예: 11KG)
            qty_m2 = re.search(r"(\d+(?:\.\d+)?\s*[A-Z가-힣]+)", rest, re.IGNORECASE)
            if qty_m2:
                qty_text = qty_m2.group(1)
                name_part = rest[: qty_m2.start()].strip()
            else:
                continue

        # 품목명 정리 (맨 앞 토큰만 쓰는 대신, 공백 제거 후 전체를 씀)
        product = name_part.strip()
        product = re.sub(r"\s{2,}", " ", product)

        rows.append({
            "photo_product": product,
            "qty_text": qty_text,
            "qty": parse_qty(qty_text),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    # 동일 품목 합산
    df["photo_product_norm"] = df["photo_product"].map(normalize_name)
    agg = (
        df.groupby(["photo_product_norm", "photo_product"], as_index=False)["qty"]
          .sum()
          .sort_values("photo_product")
    )
    agg["qty"] = agg["qty"].astype(float)
    return agg[["photo_product", "qty"]]


def get_excel_products(ws, max_consecutive_blank: int = 200):
    """
    엑셀 시트의 '상품명' 컬럼(PRODUCT_COL)에서 실제 상품 목록을 빠르게 추출.
    빈칸이 연속으로 일정 개수 나오면 더 이상 데이터가 없다고 보고 중단합니다.
    """
    products = []
    blanks = 0
    max_row = ws.max_row  # ⚠️ 반드시 한 번만 계산

    for r in range(FIRST_DATA_ROW, max_row + 1):
        v = ws.cell(r, PRODUCT_COL).value
        s = "" if v is None else str(v).strip()

        if not s:
            blanks += 1
            if blanks >= max_consecutive_blank:
                break
            continue

        blanks = 0
        products.append(s)

    # 중복 제거(순서 유지)
    seen = set()
    uniq = []
    for p in products:
        if p not in seen:
            uniq.append(p)
            seen.add(p)
    return uniq


def find_product_row(ws, excel_product: str, max_consecutive_blank: int = 200):
    blanks = 0
    max_row = ws.max_row  # ⚠️ 반드시 한 번만 계산

    for r in range(FIRST_DATA_ROW, max_row + 1):
        v = ws.cell(r, PRODUCT_COL).value
        s = "" if v is None else str(v).strip()

        if not s:
            blanks += 1
            if blanks >= max_consecutive_blank:
                break
            continue

        blanks = 0
        if s == excel_product:
            return r

    return None


def append_new_product_row(ws, product_name: str) -> int:
    """
    품목이 없으면 맨 아래에 추가 + 수식(J,N,P) 자동 생성
    (기존 스타일은 바로 위 행에서 복사)
    """
    # 마지막 데이터 행 찾기
    last = ws.max_row
    while last >= FIRST_DATA_ROW and ws.cell(last, PRODUCT_COL).value is None:
        last -= 1
    new_r = last + 1

    # 스타일 복사(위 행 기준)
    base_r = max(FIRST_DATA_ROW, new_r - 1)
    for c in range(1, ws.max_column + 1):
        ws.cell(new_r, c).value = None
        ws.cell(new_r, c)._style = copy(ws.cell(base_r, c)._style)

    ws.cell(new_r, PRODUCT_COL).value = product_name

    # 기본값
    ws.cell(new_r, 8).value = 0  # H 재고
    ws.cell(new_r, 9).value = 0  # I 입고
    ws.cell(new_r, 11).value = 0 # K 1차
    ws.cell(new_r, 12).value = 0 # L 2차
    ws.cell(new_r, 13).value = 0 # M 3차

    # 수식 (12.14 시트 구조 기준)
    ws.cell(new_r, 10).value = f"=H{new_r}+I{new_r}"               # J 보유수량
    ws.cell(new_r, 14).value = f"=K{new_r}+L{new_r}+M{new_r}"      # N 주문수량
    ws.cell(new_r, 16).value = f"=J{new_r}-N{new_r}"               # P 남은수량

    return new_r


def ensure_sheet(wb, target_name: str, template_name: str) -> None:
    if target_name in wb.sheetnames:
        return
    tpl = wb[template_name]
    new_ws = wb.copy_worksheet(tpl)
    new_ws.title = target_name

    # 새 시트는 입력열(H,I,K,L,M)을 0으로 초기화 (품목행들만)
    for r in range(FIRST_DATA_ROW, new_ws.max_row + 1):
        if new_ws.cell(r, PRODUCT_COL).value:
            for col in [8, 9, 11, 12, 13]:
                new_ws.cell(r, col).value = 0


def apply_to_excel(
    wb,
    sheet_name: str,
    category_frames: dict[str, pd.DataFrame],
    mapping: dict,
    mode: str = "덮어쓰기",
) -> None:
    ws = wb[sheet_name]

    for category, df in category_frames.items():
        if df is None or df.empty:
            continue

        col = CATEGORY_TO_COLUMN[category]

        for _, row in df.iterrows():
            photo_product = str(row["photo_product"]).strip()
            qty = float(row["qty"])

            mapped = mapping.get(normalize_name(photo_product), "")
            mapped = mapped.strip()
            if not mapped:
                # 매핑 없는 건 스킵
                continue

            r_idx = find_product_row(ws, mapped)
            if r_idx is None:
                r_idx = append_new_product_row(ws, mapped)

            if mode == "누적":
                old = ws.cell(r_idx, col).value
                old_val = float(old) if old not in (None, "") else 0.0
                ws.cell(r_idx, col).value = old_val + qty
            else:
                ws.cell(r_idx, col).value = qty


# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="재고/입고/출고(1~3차) 사진→엑셀 매핑", layout="wide")
st.title("재고/입고/1차/2차/3차 사진 업로드 → 재고파악(.xlsm) 자동 반영")

st.markdown(
    """
- 사진에서 품목/수량을 OCR로 추출(가능하면)하고, **사진 품목명 → 엑셀 품목명**을 매핑해 반영합니다.  
- OCR이 완벽하지 않을 수 있어 **표를 직접 수정**할 수 있게 해뒀습니다.
"""
)

with st.sidebar:
    st.header("1) 엑셀 업로드")
    excel_file = st.file_uploader("재고파악 .xlsm 업로드", type=["xlsm"])
    st.caption("※ 매크로 보존을 위해 .xlsm 그대로 업로드하세요.")

    st.header("2) OCR 설정(선택)")
    tesseract_cmd = st.text_input(
        "tesseract.exe 경로(비워두면 자동)",
        value="",
        placeholder=r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    )

    st.header("3) 반영 옵션")
    write_mode = st.radio("반영 방식", ["덮어쓰기", "누적"], horizontal=True)


if not excel_file:
    st.info("왼쪽 사이드바에서 재고파악(.xlsm)을 먼저 업로드하세요.")
    st.stop()

excel_bytes = excel_file.getvalue()
wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), keep_vba=True)
sheet_names = wb.sheetnames

c1, c2, c3 = st.columns([1, 1, 1])
with c1:
    template_sheet = st.selectbox("템플릿(복제 기준) 시트", sheet_names, index=len(sheet_names)-1)
with c2:
    target_sheet = st.text_input("반영할 시트명 (예: 12.15)", value=sheet_names[-1])
with c3:
    st.write("")
    st.write("")
    if st.button("시트 없으면 생성", use_container_width=True):
        ensure_sheet(wb, target_sheet, template_sheet)
        st.success(f"시트 확인 완료: {target_sheet}")

# 시트 준비
ensure_sheet(wb, target_sheet, template_sheet)
ws_target = wb[target_sheet]
excel_products = get_excel_products(ws_target)

st.divider()

mapping = load_mapping()  # {normalize(photo_name): excel_name}

st.subheader("2) 사진 업로드 (5종)")
upload_cols = st.columns(5)
uploads = {}

for i, cat in enumerate(["재고", "입고", "1차", "2차", "3차"]):
    with upload_cols[i]:
        uploads[cat] = st.file_uploader(
            f"{cat} 사진",
            type=["png", "jpg", "jpeg", "webp"],
            accept_multiple_files=True,
            key=f"u_{cat}",
        )

# OCR + 파싱
category_frames = {}
for cat, files in uploads.items():
    items = []
    if files:
        for f in files:
            img = Image.open(f).convert("RGB")
            if tesseract_cmd.strip() == "":
                cmd = None
            else:
                cmd = tesseract_cmd.strip()

            try:
                text = ocr_extract_text(img, cmd)
                df = parse_table_text_to_items(text)
            except Exception:
                df = pd.DataFrame()

            # OCR 실패/빈 결과면 “수동 입력”을 위해 빈 df 하나라도 제공
            if df is None or df.empty:
                continue
            items.append(df)

    if items:
        merged = pd.concat(items, ignore_index=True)
        # 품목 합산
        merged["photo_product_norm"] = merged["photo_product"].map(normalize_name)
        merged = merged.groupby(["photo_product_norm", "photo_product"], as_index=False)["qty"].sum()
        merged = merged.sort_values("photo_product")[["photo_product", "qty"]]
        category_frames[cat] = merged
    else:
        category_frames[cat] = pd.DataFrame(columns=["photo_product", "qty"])

st.divider()

st.subheader("3) 품목 매핑 & 수량 확인/수정")

tabs = st.tabs(["재고", "입고", "1차", "2차", "3차"])
edited_frames = {}

for tab, cat in zip(tabs, ["재고", "입고", "1차", "2차", "3차"]):
    with tab:
        df = category_frames.get(cat, pd.DataFrame(columns=["photo_product", "qty"])).copy()

        if df.empty:
            st.warning(f"{cat}: OCR로 뽑힌 데이터가 없습니다. (이미지 품질/설치 상태에 따라 그럴 수 있어요)")
            st.caption("그래도 아래에서 직접 행을 추가해 입력할 수 있습니다.")
            df = pd.DataFrame([{"photo_product": "", "qty": 0.0}], columns=["photo_product", "qty"])

        # 매핑 컬럼 추가
        mapped_list = []
        for p in df["photo_product"].tolist():
            mapped_list.append(mapping.get(normalize_name(p), ""))
        df["mapped_excel_product"] = mapped_list

        st.caption("photo_product/qty/mapped_excel_product 를 직접 수정할 수 있어요.")
        df2 = st.data_editor(
            df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "mapped_excel_product": st.column_config.SelectboxColumn(
                    "mapped_excel_product",
                    help="엑셀에 있는 품목명으로 매핑하세요 (없으면 직접 타이핑도 가능)",
                    options=[""] + excel_products,
                )
            }
        )

        # 매핑 저장 반영
        for _, r in df2.iterrows():
            p = str(r.get("photo_product", "")).strip()
            m = str(r.get("mapped_excel_product", "")).strip()
            if p and m:
                mapping[normalize_name(p)] = m

        edited_frames[cat] = df2[["photo_product", "qty"]].copy()

st.divider()

cA, cB = st.columns([1, 1])
with cA:
    if st.button("매핑 저장(product_mapping.json)", use_container_width=True):
        save_mapping(mapping)
        st.success("매핑 저장 완료! 다음 실행 때도 그대로 불러옵니다.")

with cB:
    if st.button("엑셀에 반영 후 다운로드 파일 만들기", use_container_width=True):
        # 다시 edited_frames를 반영할 때, 매핑 기준으로 적용
        # (edited_frames에는 mapped_excel_product가 없으니, mapping dict를 사용)
        apply_to_excel(
            wb=wb,
            sheet_name=target_sheet,
            category_frames=edited_frames,
            mapping=mapping,
            mode=write_mode,
        )

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_name = f"재고파악_업데이트_{target_sheet}_{ts}.xlsm"

        st.download_button(
            "다운로드(.xlsm)",
            data=out.getvalue(),
            file_name=out_name,
            mime="application/vnd.ms-excel.sheet.macroEnabled.12",
            use_container_width=True,
        )
        st.success("반영 완료! 다운로드해서 사용하세요.")

